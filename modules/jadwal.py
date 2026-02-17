import os
import sqlite3
from datetime import datetime
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.prompt import Prompt, Confirm, IntPrompt
from rich.text import Text
import json

console = Console()

class JadwalManager:
    def __init__(self, db, profile):
        self.db = db
        self.profile = profile
        self.hari_list = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    
    def display_jadwal_hari(self, hari=None):
        """Tampilkan jadwal untuk hari tertentu atau semua hari"""
        if hari:
            query = "SELECT * FROM jadwal WHERE hari = ? ORDER BY waktu_mulai"
            rows = self.db.fetch_all(query, (hari,))
        else:
            query = "SELECT * FROM jadwal ORDER BY CASE hari " + \
                    "WHEN 'Senin' THEN 1 WHEN 'Selasa' THEN 2 WHEN 'Rabu' THEN 3 " + \
                    "WHEN 'Kamis' THEN 4 WHEN 'Jumat' THEN 5 WHEN 'Sabtu' THEN 6 " + \
                    "WHEN 'Minggu' THEN 7 END, waktu_mulai"
            rows = self.db.fetch_all(query)
        
        if not rows:
            console.print(f"[yellow]Tidak ada jadwal untuk {hari if hari else 'semua hari'}[/yellow]")
            return []
        
        table = Table(title=f"📅 Jadwal {hari if hari else 'Minggu Ini'}")
        table.add_column("ID", style="cyan", justify="right")
        table.add_column("Hari", style="magenta")
        table.add_column("Mata Pelajaran", style="green")
        table.add_column("Waktu", style="yellow")
        table.add_column("Ruangan", style="blue")
        table.add_column("Pengajar", style="red")
        
        for row in rows:
            waktu = f"{row['waktu_mulai']} - {row['waktu_selesai']}"
            table.add_row(
                str(row['id']),
                row['hari'],
                row['mata_pelajaran'],
                waktu,
                row['ruangan'] or "-",
                row['pengajar'] or "-"
            )
        
        console.print(table)
        return rows
    
    def add_jadwal(self):
        """Tambah jadwal baru"""
        console.print("[bold cyan]➕ Tambah Jadwal Baru[/bold cyan]")
        
        hari = Prompt.ask(
            "Hari",
            choices=self.hari_list,
            default=self.hari_list[datetime.now().weekday()]
        )
        
        mata_pelajaran = Prompt.ask("Mata Pelajaran/Kegiatan")
        waktu_mulai = Prompt.ask("Waktu Mulai (HH:MM)", default="08:00")
        waktu_selesai = Prompt.ask("Waktu Selesai (HH:MM)", default="09:30")
        ruangan = Prompt.ask("Ruangan (opsional)", default="")
        pengajar = Prompt.ask("Pengajar (opsional)", default="")
        
        try:
            datetime.strptime(waktu_mulai, "%H:%M")
            datetime.strptime(waktu_selesai, "%H:%M")
        except ValueError:
            console.print("[red]Format waktu salah! Gunakan HH:MM[/red]")
            return
        
        self.db.execute_query('''
            INSERT INTO jadwal (hari, mata_pelajaran, waktu_mulai, waktu_selesai, ruangan, pengajar)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (hari, mata_pelajaran, waktu_mulai, waktu_selesai, ruangan, pengajar))
        
        console.print("[green]✓ Jadwal berhasil ditambahkan![/green]")
    
    def edit_jadwal(self):
        """Edit jadwal yang ada"""
        rows = self.display_jadwal_hari()
        if not rows:
            return
        
        try:
            jadwal_id = int(Prompt.ask("Masukkan ID jadwal yang akan diedit"))
        except ValueError:
            console.print("[red]ID harus angka![/red]")
            return
        
        current = self.db.fetch_one("SELECT * FROM jadwal WHERE id = ?", (jadwal_id,))
        
        if not current:
            console.print("[red]ID jadwal tidak ditemukan![/red]")
            return
        
        console.print(f"[yellow]Mengedit: {current['mata_pelajaran']}[/yellow]")
        
        hari = Prompt.ask("Hari", default=current['hari'])
        mata_pelajaran = Prompt.ask("Mata Pelajaran", default=current['mata_pelajaran'])
        waktu_mulai = Prompt.ask("Waktu Mulai", default=current['waktu_mulai'])
        waktu_selesai = Prompt.ask("Waktu Selesai", default=current['waktu_selesai'])
        ruangan = Prompt.ask("Ruangan", default=current['ruangan'] or "")
        pengajar = Prompt.ask("Pengajar", default=current['pengajar'] or "")
        
        self.db.execute_query('''
            UPDATE jadwal 
            SET hari = ?, mata_pelajaran = ?, waktu_mulai = ?, 
                waktu_selesai = ?, ruangan = ?, pengajar = ?
            WHERE id = ?
        ''', (hari, mata_pelajaran, waktu_mulai, waktu_selesai, ruangan, pengajar, jadwal_id))
        
        console.print("[green]✓ Jadwal berhasil diupdate![/green]")
    
    def delete_jadwal(self):
        """Hapus jadwal"""
        rows = self.display_jadwal_hari()
        if not rows:
            return
        
        try:
            jadwal_id = int(Prompt.ask("Masukkan ID jadwal yang akan dihapus"))
        except ValueError:
            console.print("[red]ID harus angka![/red]")
            return
        
        if Confirm.ask(f"Yakin hapus jadwal ID {jadwal_id}?"):
            self.db.execute_query("DELETE FROM jadwal WHERE id = ?", (jadwal_id,))
            console.print("[green]✓ Jadwal berhasil dihapus![/green]")
    
    def check_today_schedule(self):
        """Lihat jadwal hari ini"""
        today = self.hari_list[datetime.now().weekday()]
        console.print(f"[bold cyan]📋 Jadwal Hari Ini ({today}):[/bold cyan]")
        self.display_jadwal_hari(today)

    def export_jadwal(self):
        """Export jadwal ke JSON dan Excel"""
        rows = self.db.fetch_all("SELECT * FROM jadwal ORDER BY hari, waktu_mulai")
        
        if not rows:
            console.print("[yellow]Tidak ada jadwal untuk diexport[/yellow]")
            return
        
        # Pastikan folder exports ada
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        export_dir = os.path.join(base_dir, "exports")
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        console.print("\n[bold cyan]Pilih format export:[/bold cyan]")
        console.print("1. 📄 JSON")
        console.print("2. 📊 Excel (.xlsx)")
        console.print("3. 📦 Keduanya (JSON + Excel)")
        
        fmt = Prompt.ask("Format", choices=["1", "2", "3"], default="3")
        
        if fmt in ["1", "3"]:
            # Export JSON
            export_data = {
                "export_date": datetime.now().isoformat(),
                "total_jadwal": len(rows),
                "jadwal": rows
            }
            json_file = os.path.join(export_dir, f"jadwal_export_{timestamp}.json")
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            console.print(f"[green]✓ JSON disimpan: {json_file}[/green]")
        
        if fmt in ["2", "3"]:
            # Export Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Jadwal Pelajaran"
                
                # Header style
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                headers = ["ID", "Hari", "Mata Pelajaran", "Waktu Mulai", "Waktu Selesai", "Ruangan", "Pengajar"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                for row_idx, row in enumerate(rows, 2):
                    ws.cell(row=row_idx, column=1, value=row['id']).border = thin_border
                    ws.cell(row=row_idx, column=2, value=row['hari']).border = thin_border
                    ws.cell(row=row_idx, column=3, value=row['mata_pelajaran']).border = thin_border
                    ws.cell(row=row_idx, column=4, value=row['waktu_mulai']).border = thin_border
                    ws.cell(row=row_idx, column=5, value=row['waktu_selesai']).border = thin_border
                    ws.cell(row=row_idx, column=6, value=row['ruangan'] or "-").border = thin_border
                    ws.cell(row=row_idx, column=7, value=row['pengajar'] or "-").border = thin_border
                
                # Auto-width columns
                for col in ws.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 4
                
                excel_file = os.path.join(export_dir, f"jadwal_export_{timestamp}.xlsx")
                wb.save(excel_file)
                console.print(f"[green]✓ Excel disimpan: {excel_file}[/green]")
            except ImportError:
                console.print("[red]openpyxl belum terinstall! Jalankan: pip install openpyxl[/red]")
    
    def run(self):
        """Antarmuka utama jadwal"""
        while True:
            os.system('cls' if os.name == 'nt' else 'clear')
            console.print("\n[bold cyan]📅 Manajer Jadwal Pelajaran[/bold cyan]")
            
            menu = Table(show_header=False, box=None)
            menu.add_column("Opsi", style="cyan")
            menu.add_column("Keterangan", style="green")
            
            menu_items = [
                ("1", "👀 Lihat Semua Jadwal"),
                ("2", "📅 Lihat Jadwal Hari Ini"),
                ("3", "➕ Tambah Jadwal"),
                ("4", "✏️ Edit Jadwal"),
                ("5", "🗑️ Hapus Jadwal"),
                ("6", "📊 Export Jadwal"),
                ("0", "🔙 Kembali")
            ]
            
            for item in menu_items:
                menu.add_row(item[0], item[1])
            
            console.print(menu)
            
            choice = Prompt.ask("Pilihan", choices=["0", "1", "2", "3", "4", "5", "6"], default="0")
            
            if choice == "0":
                break
            elif choice == "1":
                self.display_jadwal_hari()
            elif choice == "2":
                self.check_today_schedule()
            elif choice == "3":
                self.add_jadwal()
            elif choice == "4":
                self.edit_jadwal()
            elif choice == "5":
                self.delete_jadwal()
            elif choice == "6":
                self.export_jadwal()
            
            if choice != "0":
                console.print("[dim]Tekan Enter untuk melanjutkan...[/dim]")
                input()